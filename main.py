import openpyxl 
import numpy as np 
import pandas as pd 
import os 
import datetime 
import time 
from glob import glob 
import PySimpleGUI as sg

def get_year_month():
    JST = datetime.timezone(datetime.timedelta(hours=+9), 'JST')
    today = str(datetime.datetime.now(JST))
    return today.split(".")[0]

def add_cell_values(ws: object, datas: list, row: int):
    '''
    セルにデータを追加する
    [
        "日付", "担当ID", "新規", "機変", "クレカ", "でんき", "+1コレクション", "ネットワーク", "他", 
    ]
    '''
    assert len(datas) == 10
    for number, data in enumerate(datas):
        ws.cell(row, number+1, data)
    return ws 

# def input_data(name): 
#     try:
#         now = get_year_month()
#         shinki = input("新規台数は？")
#         kihen = input("機変台数は？")
#         card = input("クレカは？")
#         denki = input("でんきは？")
#         collection = input("+1コレクション合計金額は？")
#         network = input("ネットワーク獲得数は？")
#         other = input("他の商材数は？")
#         num_customer = input("総接客数は？")
#         datas = [now, name, int(shinki), int(kihen), int(card), int(denki), float(collection), int(network), int(other), int(num_customer)]
#         return datas 
#     except Exception as e:
#         print("不正な値が検知されました。")
#         raise ValueError
    
    
def create_this_year_book(year: int):
    '''今年のブックを作成する'''
    files = os.listdir("成績管理表")
    flag = 0 

    for file in files:
        if str(year) in file:
            flag = 1 

    if not flag:
        wb = openpyxl.Workbook()
        wb.save(f"成績管理表/{str(year)}年分.xlsx")
        wb.close()
        
def create_this_month_sheet(year: int, month: int):
    '''今月のシートを作成する'''
    wb = openpyxl.load_workbook(f"成績管理表/{str(year)}年分.xlsx")
    files = wb.sheetnames 
    flag = 0 

    for file in files:
        if str(month) + "月" in file:
            flag = 1 

    if not flag:
        ws = wb.create_sheet(str(month) + "月")
        datas = ["日付", "担当者", "新規", "機変", "クレカ", "でんき", "+1 コレクション", "ネットワーク", "他", "接客数"]
        ws = add_cell_values(ws, datas, 1)
        wb.save(f"成績管理表/{str(year)}年分.xlsx")

    wb.close()
    

def delete_this_year_sheet(year: int):
    '''[Sheet]を除去する'''
    wb = openpyxl.load_workbook(f"成績管理表/{str(year)}年分.xlsx")
    files = wb.sheetnames 

    if "Sheet" in files:
        ws = wb["Sheet"]
        wb.remove(ws)
        wb.save(f"成績管理表/{str(year)}年分.xlsx")
    wb.close()
    
    
def main():
    os.makedirs("成績管理表", exist_ok=True)
    today = get_year_month()
    year, month = int(today.split("-")[0]), int(today.split("-")[1])
    
    # ブックとシートの作成
    create_this_year_book(year)
    create_this_month_sheet(year, month)
    delete_this_year_sheet(year)
    
    # シートを取得する
    wb = openpyxl.load_workbook(f"成績管理表/{str(year)}年分.xlsx")
    ws = wb[str(month) + "月"]
    max_row = ws.max_row + 1 
    
    # GUIの作成
    layout = [[sg.Text("担当者名", size=(10, 1)), sg.InputText(key="-NAME-", default_text="(例) tanaka", size=(14, 1))], 
             [sg.Text("新規台数", size=(10, 1)), sg.Combo(list(range(0, 11)), key="-SINKI-", default_value=0)],
             [sg.Text("機種変更", size=(10, 1)), sg.Combo(list(range(0, 21)), key="-KIHEN-", default_value=0)],
             [sg.Text("クレカ", size=(10, 1)), sg.Combo(list(range(0, 11)), key="-CARD-", default_value=0)],
             [sg.Text("でんき", size=(10, 1)), sg.Combo(list(range(0, 11)), key="-DENKI-", default_value=0)],
             [sg.Text("＋１", size=(10, 1)), sg.InputText(key="-COLLECT-", default_text="(例) 12000", size=(14, 1))],
             [sg.Text("NW", size=(10, 1)), sg.Combo(list(range(0, 11)), key="-NW-", default_value=0)], 
             [sg.Text("他", size=(10, 1)), sg.Combo(list(range(0, 11)), key="-OTHER-", default_value=0)], 
             [sg.Text("総接客数", size=(10, 1)), sg.Combo(list(range(0, 30)), key="-CUSTOMER-", default_value=5)], 
             [sg.Output(size=(80, 3))], 
             [sg.Button("実行", key="-SUBMIT-")]]
    window = sg.Window("売上管理アプリ", layout, size=(400, 400))
    
    while True:
        event, values = window.read()
        
        if event == sg.WIN_CLOSED:
            break 
            
        if event == "-SUBMIT-":
            name = values["-NAME-"]
            sinki = values["-SINKI-"]
            kihen = values["-KIHEN-"]
            card = values["-CARD-"]
            denki = values["-DENKI-"]
            collection = values["-COLLECT-"]
            nw = values["-NW-"]
            other = values["-OTHER-"]
            n_customer = values["-CUSTOMER-"]
            
            if(name == "" or name == "(例) tanaka" or collection == "(例) 12000" or collection == ""):
                print("不正な入力が検知されました、もう一度やり直してください。")
                time.sleep(2)
                print("")
                
            else:
                try:
                    data = [today, name, int(sinki), int(kihen), int(card), int(denki), int(collection), int(nw), int(other), int(n_customer)]
                    add_cell_values(ws, data, max_row)
                    wb.save(f"成績管理表/{str(year)}年分.xlsx")
                    wb.close()
                    print(f"正しく反映されました。{name}様、本日も勤務お疲れさまでした。")
                    time.sleep(2)
                    break 
                except Exception as e:
                    print("予期せぬエラーが発生しました。")
                    time.sleep(2)
                    break 
    window.close()
    
    
if __name__ == "__main__":
    main()