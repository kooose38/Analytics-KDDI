import openpyxl 
from openpyxl.chart import PieChart, Reference, LineChart, BarChart
import pandas as pd 
import numpy as np 
import os 
import PySimpleGUI as sg 
import time 

COLLECTION_THRESHOLD = 10000
SHEET_NAMES = ["獲得数", "獲得推移", "１接客当たりの獲得数", "付帯率"]

def calc_product_sum(df):
    return df[["新規", "機変", "クレカ", "でんき", "ネットワーク", "+1 コレクション", "他"]].sum().to_frame().rename(columns={0: "今月の合計獲得数"})

def calc_product_transit(df):
    _df = df.copy()
    _df["月日"] = _df["日付"].apply(lambda x: x.split(" ")[0])
    return _df.groupby("月日").sum()

def calc_per_customer_service(df, name="店舗"):
    sum_customer = df["接客数"].sum()
    return df[["新規", "機変", "クレカ", "でんき", "ネットワーク", "+1 コレクション", "他"]].sum().to_frame().rename(columns={0: f"{name}-１接客当たりの獲得数"}) / sum_customer

def calc_incidental_rate(df, name="店舗"):
    sum_uriage = df["新規"].sum() + df["機変"].sum()
    return df[["クレカ", "でんき", "ネットワーク", "+1 コレクション", "他"]].sum().to_frame().rename(columns={0: f"{name}-付帯率"}) / sum_uriage


def saved_excel(df1, df2, df3, df4, sheet_names, author, year, month):
    df1.to_excel(f"分析表/{author}-{str(year)}年-{str(month)}月.xlsx", sheet_name=sheet_names[0])
    with pd.ExcelWriter(f"分析表/{author}-{str(year)}年-{str(month)}月.xlsx", mode="a") as w:
        df2.to_excel(w, sheet_name=sheet_names[1])
        df3.to_excel(w, sheet_name=sheet_names[2])
        df4.to_excel(w, sheet_name=sheet_names[3])
        
        
def plot_pie(wb, sheet_name):
    ws = wb[sheet_name]
    pie = PieChart()
    data = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=7)
    label = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=7)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(label)
    pie.title = sheet_name
    return pie 

def plot_line(wb, sheet_name):
    ws = wb[sheet_name]
    max_row = ws.max_row + 1 
    line = LineChart()
    data = Reference(ws, min_col=2, max_col=8, min_row=1, max_row=max_row)
    label = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=max_row)
    line.add_data(data, titles_from_data=True)
    line.set_categories(label)
    line.title = sheet_name 
    return line 
    
def plot_bar(wb, sheet_name):
    ws = wb[sheet_name]
    max_row = ws.max_row + 1 
    bar = BarChart()
    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=max_row)
    label = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=max_row)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(label)
    bar.title = sheet_name
    return bar 

def get_year():
    files = os.listdir("成績管理表")
    arr = []
    for file in files:
        if ".ipynb_checkpoints" in file:
            continue
        arr.append(int(file.split("年")[0][-4:]))
    return arr 

def main():
    os.makedirs("分析表", exist_ok=True)
    years = get_year()
    layout = [[sg.Text("過去のデータから分析グラフを作成します。")],
             [sg.Combo(years, key="-YEAR-", default_value=years[len(years)-1]), sg.Text("年")], 
             [sg.Combo(list(range(1, 13)), key="-MONTH-", default_value=1), sg.Text("月")], 
              [sg.Output(size=(80, 5))], 
             [sg.Button("実行", key="-SUBMIT-")]]

    window = sg.Window("売上分析アプリ", layout, size=(400, 300))

    while True:
        event, values = window.read()

        if event == sg.WIN_CLOSED:
            break 

        if event == "-SUBMIT-":
            try:
                year = int(values["-YEAR-"])
                month = int(values["-MONTH-"])
            
                df = pd.read_excel(f"成績管理表/{str(year)}年分.xlsx", sheet_name=str(month) + "月")
                df["+1 コレクション"] = df["+1 コレクション"] / COLLECTION_THRESHOLD

                for author in df["担当者"].unique():
                    df_author = df[df["担当者"] == author]
                    df_author_sum = calc_product_sum(df_author)
                    df_author_transit = calc_product_transit(df_author)
                    df_per_customer = calc_per_customer_service(df_author, name=author)
                    df_all_customer = calc_per_customer_service(df)
                    df_per_customer = pd.merge(df_per_customer, df_all_customer, how="outer", left_index=True, right_index=True)
                    df_incidental_rate = calc_incidental_rate(df_author, name=author)
                    df_all_incidental_rate = calc_incidental_rate(df)
                    df_incidental_rate = pd.merge(df_incidental_rate, df_all_incidental_rate, how="outer", left_index=True, right_index=True)

                    saved_excel(df_author_sum, df_author_transit, df_per_customer, df_incidental_rate, SHEET_NAMES, author, year, month)

                    wb = openpyxl.load_workbook(f"分析表/{author}-{str(year)}年-{str(month)}月.xlsx")
                    pie = plot_pie(wb, SHEET_NAMES[0])
                    line = plot_line(wb, SHEET_NAMES[1])
                    bar1 = plot_bar(wb, SHEET_NAMES[2])
                    bar2 = plot_bar(wb, SHEET_NAMES[3])

                    ws_plot = wb.create_sheet("分析表グラフ")

                    ws_plot.add_chart(pie, "A1")
                    ws_plot.add_chart(line, "A20")
                    ws_plot.add_chart(bar1, "A40")
                    ws_plot.add_chart(bar2, "A60")

                    wb.save(f"分析表/{author}-{str(year)}年-{str(month)}月.xlsx")

                print("正常に完了しました。分析表フォルダーより担当者別にダウンロードしてください。")
                time.sleep(5)
                break 

            except Exception as e:
                print("予期せぬエラーが発生しました。", e)
                time.sleep(2)
                break 
        

if __name__ == "__main__":
    main()